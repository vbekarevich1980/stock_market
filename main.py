import os
import time
import datetime

from pathlib import Path
from openpyxl.drawing.image import Image
import openpyxl

from twisted.internet import defer # reactor,
from scrapy.crawler import CrawlerRunner, CrawlerProcess
from scrapy.utils.log import configure_logging
from scrapy.utils.project import get_project_settings
from stock_market.spiders.stock_market_spider import StockMarketSpider
from stock_market.spiders.companies_spider import CompaniesSpider


from scrapy.utils.reactor import install_reactor

if __name__ == "__main__":

    #os.system('docker pull scrapinghub/splash')
    # time.sleep(10)
    #os.system('docker run -d -p 8050:8050 --rm --name splash scrapinghub/splash')
    #time.sleep(10)

    configure_logging()
    settings = get_project_settings()
    # runner = CrawlerRunner(settings)
    #
    # @defer.inlineCallbacks
    # def crawl():
    #     yield runner.crawl(ProxySpider)
    #     #yield runner.crawl(BroswaySpider)
    #     reactor.stop()
    #
    # crawl()
    # reactor.run()


    install_reactor('twisted.internet.asyncioreactor.AsyncioSelectorReactor')
    process = CrawlerProcess(settings=settings)
    #process.crawl(CompaniesSpider)
    process.crawl(StockMarketSpider)
    process.start()

    # Extract the tickers from the first column
    destination_xlsx_file = Path('Stock Market.xlsx')
    destination_wb_obj = openpyxl.load_workbook(destination_xlsx_file)
    destination_sheet = destination_wb_obj.active

    src_xlsx_file = Path('stock.xlsx')
    src_wb_obj = openpyxl.load_workbook(src_xlsx_file)
    src_sheet = src_wb_obj.active

    chart_column_dict = {"revenue": "AN", "net_income": "AO", "esp": "AP",
        "price_sales": "AQ", "pe_ratio": "AR", "price_book": "AS",
        "market_cap": "AT", }

    for destination_row, src_row in zip(
            destination_sheet.iter_rows(min_row=2, values_only=False),
            src_sheet.iter_rows(min_row=2, values_only=False)):
        for destination_cell, src_cell in zip(destination_row, src_row):
            destination_cell.value = src_cell.value

        for chart_type, column_id in chart_column_dict.items():
            # Load the images and add to worksheet and anchor next to cells
            img_filename = os.path.join('stock_market', 'screenshots',
                                        f'{src_row[0].value}_{chart_type}.png')
            img = Image(img_filename)
            img.object_position = 1
            img.width = 476
            img.height = 357
            destination_sheet.add_image(img, f'{column_id}{src_row[0].row}')

    destination_wb_obj.save(f'Stock Market {datetime.datetime.today()}.xlsx')
    destination_wb_obj.close()

    #os.system('docker stop splash')
